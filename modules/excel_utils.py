import openpyxl
from openpyxl.styles import numbers



import openpyxl
from openpyxl.styles import numbers


def copy_data_to_excel(data, template_filepath, sheet_name, header_row=1):  # 添加 header_row 参数
    """将数据复制到 Excel 文件的指定工作表，保留表头。"""
    try:
        template_wb = openpyxl.load_workbook(template_filepath)
        template_sheet = template_wb[sheet_name]

        # 查找第一个空行
        row_index = header_row + 1  # 从表头下一行开始写入数据
        while template_sheet.cell(row=row_index, column=1).value is not None:
            row_index += 1

        for row_data in data:
            for col_index, cell_data in enumerate(row_data, start=1):
                cell = template_sheet.cell(row=row_index, column=col_index)
                cell.value = cell_data
                cell.number_format = numbers.FORMAT_TEXT
            row_index += 1

        template_wb.save(template_filepath)
        return True
    except FileNotFoundError:
        print(f"文件 {template_filepath} 未找到。")
        return False
    except KeyError:
        print(f"工作表 {sheet_name} 未找到。")
        return False
    except Exception as e:
        print(f"复制数据到 Excel 时发生错误：{e}")
        import traceback
        traceback.print_exc()
        return False

# def copy_data_to_excel(data, template_filepath, sheet_name, header_row=1):
#     """将数据复制到 Excel 文件的指定工作表，保留表头，并清空原有数据。"""
#     try:
#         template_wb = openpyxl.load_workbook(template_filepath)
#         template_sheet = template_wb[sheet_name]
#
#         # 清空原有数据，保留表头
#         max_row = template_sheet.max_row  # 获取最大行号
#         if max_row > header_row:  # 如果有数据超出表头行
#             template_sheet.delete_rows(header_row + 1, max_row - header_row)  # 删除表头行之后的所有行
#
#         # 写入数据
#         row_index = header_row + 1  # 从表头下一行开始写入数据
#         for row_data in data:
#             for col_index, cell_data in enumerate(row_data, start=1):
#                 cell = template_sheet.cell(row=row_index, column=col_index)
#                 cell.value = cell_data
#                 cell.number_format = numbers.FORMAT_TEXT
#             row_index += 1
#
#         template_wb.save(template_filepath)
#         return True
#     except FileNotFoundError:
#         print(f"文件 {template_filepath} 未找到。")
#         return False
#     except KeyError:
#         print(f"工作表 {sheet_name} 未找到。")
#         return False
#     except Exception as e:
#         print(f"复制数据到 Excel 时发生错误：{e}")
#         import traceback
#         traceback.print_exc()
#         return False
#
# def copy_data_to_excel(data, template_filepath, sheet_name):
#     """将数据复制到 Excel 文件的指定工作表，覆盖原有数据。"""
#     try:
#         template_wb = openpyxl.load_workbook(template_filepath)
#         template_sheet = template_wb[sheet_name]
#
#         for row_index, row_data in enumerate(data, start=2):
#             for col_index, cell_data in enumerate(row_data, start=1):
#                 cell = template_sheet.cell(row=row_index, column=col_index)
#                 cell.value = cell_data
#                 cell.number_format = numbers.FORMAT_TEXT  # 强制设置为文本格式
#
#         template_wb.save(template_filepath)
#         return True
#     except FileNotFoundError:
#         print(f"文件 {template_filepath} 未找到。")
#         return False
#     except KeyError:
#         print(f"工作表 {sheet_name} 未找到。")
#         return False
#     except Exception as e:
#         print(f"复制数据到 Excel 时发生错误：{e}")
#         import traceback
#         traceback.print_exc()
#         return False



def copy_data_to_excel_with_header(data, filename, sheetname, header_name):
    """将数据复制到 Excel 文件的指定工作表，包含表头。"""
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]

        # 清空数据，包括表头
        # if sheet.max_row >= 1:
        #     sheet.delete_rows(1, sheet.max_row)

        sheet.cell(row=1, column=1).value = header_name  # 写入表头

        if data:
            for i, value in enumerate(data):
                sheet.cell(row=i + 2, column=1).value = value

        workbook.save(filename)
        return True
    except FileNotFoundError:
        print(f"文件 {filename} 未找到。")
        return False
    except KeyError:
        print(f"工作表 {sheetname} 未找到。")
        return False
    except Exception as e:
        print(f"复制数据到 Excel 时发生错误：{e}")
        return False


def copy_data_to_excel_with_header_split_columns(data, template_filepath, sheet_name, header_name, max_rows_per_column):
    """将数据分列复制到 Excel，包含表头，并清空原有数据（保留表头）。"""
    try:
        wb = openpyxl.load_workbook(template_filepath)
        sheet = wb[sheet_name]

        # 清空除表头外的数据
        if sheet.max_row > 1:  # 确保有数据需要删除
            sheet.delete_rows(2, sheet.max_row)  # 从第二行开始删除，保留表头

        sheet.cell(row=1, column=1).value = header_name  # 写入表头

        for i, item in enumerate(data):
            row_index = (i % max_rows_per_column) + 1
            col_index = (i // max_rows_per_column) + 1
            cell = sheet.cell(row=row_index, column=col_index) # 获取单元格对象
            cell.value = item
            cell.number_format = numbers.FORMAT_TEXT # 设置单元格格式为文本

        wb.save(template_filepath)
        return True
    except FileNotFoundError:
        print(f"文件 {template_filepath} 未找到。")
        return False
    except KeyError:
        print(f"工作表 {sheet_name} 未找到。")
        return False
    except Exception as e:
        print(f"复制数据到 Excel 时发生错误：{e}")
        import traceback
        traceback.print_exc()
        return False

#
# def copy_business_numbers_to_excel_with_header_split_columns(data, template_filepath, sheet_name, header_name, max_rows_per_column):
#     """将数据分列复制到 Excel，包含表头，并清空原有数据（保留表头）。"""
#     try:
#         wb = openpyxl.load_workbook(template_filepath)
#         sheet = wb[sheet_name]
#
#         # 清空除表头外的数据
# #@        if sheet.max_row > 1:  # 确保有数据需要删除=============================================================================
# #@            sheet.delete_rows(min_row=2, max_row=sheet.max_row)  # 正确的新版本用法，**保留这行** 从第二行开始删除，保留表头
#
#         sheet.cell(row=1, column=1).value = header_name
#
#         num_cols_needed = (len(data) + max_rows_per_column - 1) // max_rows_per_column
#
#         for i, item in enumerate(data):
#             row_index = (i % max_rows_per_column) + 1
#             col_index = (i // max_rows_per_column) + 1
#             sheet.cell(row=row_index, column=col_index).value = item
#
#         wb.save(template_filepath)
#         return True
#     except Exception as e:
#         print(f"复制数据到 Excel 时发生错误：{e}")
#         import traceback
#         traceback.print_exc()
#         return False


# def copy_business_numbers_to_excel_with_header_split_columns(data, template_filepath, sheet_name, header_name, max_rows_per_column):
#     """将数据分列复制到 Excel，包含表头，并清空原有数据（保留表头）。"""
#     try:
#         wb = openpyxl.load_workbook(template_filepath)
#         sheet = wb[sheet_name]
#
#         # 清空除表头外的数据
#         if sheet.max_row > 1:  # 确保有数据需要删除
#             sheet.delete_rows(min_row=2, max_row=sheet.max_row)  # 从第二行开始删除，保留表头
#
#         sheet.cell(row=1, column=1).value = header_name  # 写入表头（如果需要重新写入）
#
#         num_cols_needed = (len(data) + max_rows_per_column - 1) // max_rows_per_column
#
#         for i, item in enumerate(data):
#             row_index = (i % max_rows_per_column) + 1  # 加 1 是因为第一行是表头
#             col_index = (i // max_rows_per_column) + 1  # 加 1 是因为第一列是表头
#             sheet.cell(row=row_index, column=col_index).value = item
#
#         wb.save(template_filepath)
#         return True
#     except Exception as e:
#         print(f"复制数据到 Excel 时发生错误：{e}")
#         import traceback
#         traceback.print_exc() # 打印详细错误信息，方便调试
#         return False


# def copy_business_numbers_to_excel_with_header_split_columns(data, template_filepath, sheet_name, header_name, max_rows_per_column):
#     """将数据分列复制到 Excel，包含表头。"""
#     try:
#         wb = openpyxl.load_workbook(template_filepath)
#         sheet = wb[sheet_name]
#
#         sheet.cell(row=1, column=1).value = header_name  # 写入表头
#
#         num_cols_needed = (len(data) + max_rows_per_column - 1) // max_rows_per_column
#
#         for i, item in enumerate(data):
#             row_index = (i % max_rows_per_column) + 1  # 加 1 是因为第一行是表头
#             col_index = (i // max_rows_per_column) + 1  # 加 1 是因为第一列是表头
#             sheet.cell(row=row_index, column=col_index).value = item
#
#         wb.save(template_filepath)
#         return True
#     except Exception as e:
#         print(f"复制数据到 Excel 时发生错误：{e}")
#         return False



# import openpyxl
# from openpyxl.styles import numbers
#
# def copy_data_to_excel(data, template_filepath, sheet_name): # 正确的函数定义，接收 template_filepath 和 sheet_name 参数
#     """将数据复制到 Excel 文件的指定工作表，覆盖原有数据。"""
#
#     try:
#         template_wb = openpyxl.load_workbook(template_filepath) # 使用传递进来的 template_filepath
#         template_sheet = template_wb[sheet_name] # 使用传递进来的 sheet_name
#
#         for row_index, row_data in enumerate(data, start=2):
#             for col_index, cell_data in enumerate(row_data, start=1):
#                 cell = template_sheet.cell(row=row_index, column=col_index)
#                 cell.value = cell_data
#                 cell.number_format = numbers.FORMAT_TEXT  # 强制设置为文本格式
#
#         template_wb.save(template_filepath)
#         return True
#     except FileNotFoundError:
#         print(f"文件 {template_filepath} 未找到。") # 打印正确的错误信息
#         return False
#     except KeyError:
#         print(f"工作表 {sheet_name} 未找到。") # 打印正确的错误信息
#         return False
#     except Exception as e:
#         print(f"复制数据到 Excel 时发生错误：{e}")
#         import traceback
#         traceback.print_exc()
#         return False
#
# def copy_data_to_excel_with_header(data, filename, sheetname, header_name):
#     """将数据复制到 Excel 文件的指定工作表，包含表头。"""
#     try:
#         workbook = openpyxl.load_workbook(filename)
#         sheet = workbook[sheetname]
#
#         # 清空数据，包括表头
#         if sheet.max_row >= 1:
#             sheet.delete_rows(1, sheet.max_row)
#
#         sheet.cell(row=1, column=1).value = header_name  # 写入表头
#
#         if data:
#             for i, value in enumerate(data):
#                 sheet.cell(row=i + 2, column=1).value = value
#
#         workbook.save(filename)
#         return True
#     except FileNotFoundError:
#         print(f"文件 {filename} 未找到。")
#         return False
#     except KeyError:
#         print(f"工作表 {sheetname} 未找到。")
#         return False
#     except Exception as e:
#         print(f"复制数据到 Excel 时发生错误：{e}")
#         return False

'''
def copy_data_to_excel(data, filename, sheetname):
    """将数据复制到 Excel 文件的指定工作表，覆盖原有数据（兼容旧版本 openpyxl）。"""

    try:
        template_wb = openpyxl.load_workbook(template_filepath)
        template_sheet = template_wb[sheet_name]

        for row_index, row_data in enumerate(data, start=2):
            for col_index, cell_data in enumerate(row_data, start=1):
                cell = template_sheet.cell(row=row_index, column=col_index)
                cell.value = cell_data
                cell.number_format = numbers.FORMAT_TEXT # 强制设置为文本格式

        template_wb.save(template_filepath)
        return True
    except Exception as e:
        print(f"写入Excel时发生错误：{e}")
        import traceback
        traceback.print_exc()
        return False
    # try:
    #     workbook = openpyxl.load_workbook(filename)
    #     sheet = workbook[sheetname]
    #
    #     if sheet.max_row > 1:  # 确保有数据需要删除
    #         sheet.delete_rows(2, sheet.max_row)  # 旧版本 delete_rows 的用法
    #
    #     if data:
    #         for row_index, row_data in enumerate(data, start=2):
    #             for col_index, cell_value in enumerate(row_data, start=1):
    #                 sheet.cell(row=row_index, column=col_index, value=cell_value)
    #
    #     workbook.save(filename)
    #     return True
    # except FileNotFoundError:
    #     print(f"文件 {filename} 未找到。")
    #     return False
    # except KeyError:
    #     print(f"工作表 {sheetname} 未找到。")
    #     return False
    # except Exception as e:
    #     print(f"复制数据到 Excel 时发生错误：{e}")
    #     return False


def copy_data_to_excel_with_header(data, filename, sheetname, header_name):  # 新增一个函数，处理表头
    """将数据复制到 Excel 文件的指定工作表，包含表头。(兼容旧版本openpyxl)"""
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]

        # 清空数据，包括表头 (兼容旧版本)
        if sheet.max_row >= 1: #确保有数据需要删除，包括表头
            sheet.delete_rows(1, sheet.max_row)  # 旧版本 delete_rows 的用法，从第一行开始删除

        sheet.cell(row=1, column=1).value = header_name  # 写入表头

        if data:
            for i, value in enumerate(data):
                sheet.cell(row=i + 2, column=1).value = value

        workbook.save(filename)
        return True
    except FileNotFoundError:
        print(f"文件 {filename} 未找到。")
        return False
    except KeyError:
        print(f"工作表 {sheetname} 未找到。")
        return False
    except Exception as e:
        print(f"复制数据到 Excel 时发生错误：{e}")
        return False

'''


#import openpyxl

'''
def copy_data_to_excel(data, filename, sheetname):
    """将数据复制到 Excel 文件的指定工作表，覆盖原有数据（使用 openpyxl 正确清空数据）。
       此版本不处理表头。data 必须是列表的列表。"""
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]

        # 清空除表头外的数据
        sheet.delete_rows(min_row=2, max_row=sheet.max_row)

        # 写入数据
        if data:
            for row_index, row_data in enumerate(data, start=2):
                for col_index, cell_value in enumerate(row_data, start=1):
                    sheet.cell(row=row_index, column=col_index, value=cell_value)

        workbook.save(filename)
        return True
    except FileNotFoundError:
        print(f"文件 {filename} 未找到。")
        return False
    except KeyError:
        print(f"工作表 {sheetname} 未找到。")
        return False
    except Exception as e:
        print(f"复制数据到 Excel 时发生错误：{e}")
        return False

'''

'''

import openpyxl

def copy_data_to_excel(data, filename, sheetname):
    """将数据复制到 Excel 文件的指定工作表，覆盖原有数据（兼容旧版本 openpyxl）。"""
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]

        # 清空除表头外的数据 (兼容旧版本)
        if sheet.max_row > 1:  # 确保有数据需要删除
            sheet.delete_rows(2, sheet.max_row)  # 旧版本 delete_rows 的用法

        # 写入数据
        if data:
            for row_index, row_data in enumerate(data, start=2):
                for col_index, cell_value in enumerate(row_data, start=1):
                    sheet.cell(row=row_index, column=col_index, value=cell_value)

        workbook.save(filename)
        return True
    except FileNotFoundError:
        print(f"文件 {filename} 未找到。")
        return False
    except KeyError:
        print(f"工作表 {sheetname} 未找到。")
        return False
    except Exception as e:
        print(f"复制数据到 Excel 时发生错误：{e}")
        return False


def copy_data_to_excel_with_header(data, filename, sheetname, header_name): # 新增一个函数，处理表头
    """将数据复制到 Excel 文件的指定工作表，包含表头。"""
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]

        # 清空数据，包括表头
        sheet.delete_rows(min_row=1, max_row=sheet.max_row)

        sheet.cell(row=1, column=1).value = header_name  # 写入表头

        if data:
            for i, value in enumerate(data):
                sheet.cell(row=i + 2, column=1).value = value

        workbook.save(filename)
        return True
    except FileNotFoundError:
        print(f"文件 {filename} 未找到。")
        return False
    except KeyError:
        print(f"工作表 {sheetname} 未找到。")
        return False
    except Exception as e:
        print(f"复制数据到 Excel 时发生错误：{e}")
        return False
'''



