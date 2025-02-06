'''
程序名称：终端出库报表
版本 ： V4.0
功能： 处理日常的终端数据，包括复制指定数据清单，筛选指定数据，插入指定字段，匹配指定数据
开发人员： 卢鹤斌
优化人员： 卢鹤斌
注意事项 ： 待处理的文件后缀需为 .xlsx 格式 !!!!
主程序 入口 ： main.py 文件

软件使用说明：
   1、该程序所有用到的文件后缀 ， 最好为  .xlsx 格式 ，如遇 .csv格式 请转为 .xlsx格式
   2、源文件和目标文件路径 问题： 先输入“文件路径” ，在输入 “文件名称” ， 最后输入“文件子表名称”（目标文件的子表名称可以自定义）
      dwd_hzluheb_acc_sn_final_pg 文件 问题 ： 先输入“文件路径\文件名称” ，最后在输入“文件子表名称”
   3、除了 第二点 说到 的路径是分步骤输入外，其他文件的输入请直接 输入 文件路径\文件名称

'''

import openpyxl

def copy_data_to_excel(data, filename, sheetname):
    """将数据复制到 Excel 文件的指定工作表，覆盖原有数据（兼容旧版本 openpyxl）。"""
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]

        # if sheet.max_row > 1:
        #     sheet.delete_rows(2, sheet.max_row)==================================================

        if data:
            for row_index, row_data in enumerate(data, start=2):
                for col_index, cell_value in enumerate(row_data, start=1):
                    sheet.cell(row=row_index, column=col_index, value=cell_value)

        workbook.save(filename)
        return True
    except FileNotFoundError:
        print(f"文件 {filename} 未找到。")
        return False
    except KeyError:
        print(f"工作表 {sheetname} 未找到。")
        return False
    except Exception as e:
        print(f"复制数据到 Excel 时发生错误：{e}")
        return False

# 创建一个测试 Excel 文件
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"
ws.append(["Header1", "Header2"])  # 添加表头
wb.save("test_excel.xlsx")

# 测试数据
test_data = [["Data1", "Data2"], ["Data3", "Data4"]]

# 调用 copy_data_to_excel 函数进行测试
if copy_data_to_excel(test_data, "test_excel.xlsx", "Sheet1"):
    print("复制数据成功！")
else:
    print("复制数据失败！")

print(f"openpyxl 版本：{openpyxl.__version__}")