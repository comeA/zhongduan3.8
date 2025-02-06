'''
程序名称：终端出库报表
版本 ： V4.0
功能： 处理日常的终端数据，包括复制指定数据清单，筛选指定数据，插入指定字段，匹配指定数据
开发人员： 卢鹤斌
优化人员： 卢鹤斌
注意事项 ： 待处理的文件后缀需为 .xlsx 格式 !!!!
主程序 入口 ： main.py 文件

软件使用说明：
   1、该程序所有用到的文件后缀 ， 最好为  .xlsx 格式 ，如遇 .csv格式 请转为 .xlsx格式
   2、源文件和目标文件路径 问题： 先输入“文件路径” ，在输入 “文件名称” ， 最后输入“文件子表名称”（目标文件的子表名称可以自定义）
      dwd_hzluheb_acc_sn_final_pg 文件 问题 ： 先输入“文件路径\文件名称” ，最后在输入“文件子表名称”
   3、除了 第二点 说到 的路径是分步骤输入外，其他文件的输入请直接 输入 文件路径\文件名称

'''

import openpyxl
#from modules import excel_utils

import modules.excel_utils as excel_utils

def copy_business_numbers_to_template(df, template_filepath):
    """将筛选后的“业务号码”复制到模板文件，分列处理，保留原有表头。接入号复制到导入模板"""
    try:
        # 1. 清理 ' 并转换为字符串列表
        df['业务号码'] = df['业务号码'].astype(str).str.replace("^'", "", regex=True)
        business_numbers = df['业务号码'].tolist()

        print(f"business_numbers 类型：{type(business_numbers)}")
        print(f"business_numbers 内容（前 10 个）：{business_numbers[:10]}")

        # 直接调用新的excel_utils函数处理分列
        header_name = "业务号码"
        if not excel_utils.copy_business_numbers_to_excel_with_header_split_columns(business_numbers, template_filepath, "文件名称", header_name, max_rows_per_column=5000):
            print("复制业务号码失败")
            return False

        print("业务号码已成功复制到 导入模板.xlsx")
        return True

    except KeyError:
        print("错误：筛选后的数据中缺少 '业务号码' 列。")
        return False
    except Exception as e:
        print(f"处理业务号码或写入“导入模板.xlsx”文件时发生错误：{e}")
        return False


# def copy_business_numbers_to_template(df, template_filepath):
#     """将筛选后的“业务号码”复制到模板文件，分列处理，保留原有表头。"""
#     try:
#         # 1. 清理 ' 并转换为字符串列表
#         df['业务号码'] = df['业务号码'].astype(str).str.replace("^'", "", regex=True)
#         business_numbers = df['业务号码'].tolist()
#
#         print(f"business_numbers 类型：{type(business_numbers)}")
#         print(f"business_numbers 内容（前 10 个）：{business_numbers[:10]}")
#
#         # 直接调用新的excel_utils函数处理分列
#         header_name = "业务号码"
#         if not excel_utils.copy_business_numbers_to_excel_with_header_split_columns(business_numbers, template_filepath, "文件名称", header_name, max_rows_per_column=5000):
#             print("复制业务号码失败")
#             return False
#
#         print("业务号码已成功复制到 导入模板.xlsx")
#         return True
#
#     except KeyError:
#         print("错误：筛选后的数据中缺少 '业务号码' 列。")
#         return False
#     except Exception as e:
#         print(f"处理业务号码或写入“导入模板.xlsx”文件时发生错误：{e}")
#         return False


def copy_business_numbers_to_excel_with_header_split_columns(data, template_filepath, sheet_name, header_name, max_rows_per_column):
    """将数据分列复制到 Excel，包含表头。"""
    try:
        wb = openpyxl.load_workbook(template_filepath)
        sheet = wb[sheet_name]

        sheet.cell(row=1, column=1).value = header_name  # 写入表头

        num_cols_needed = (len(data) + max_rows_per_column - 1) // max_rows_per_column

        for i, item in enumerate(data):
            row_index = (i % max_rows_per_column) + 1 #加1是因为第一行是表头
            col_index = (i // max_rows_per_column) + 1 #加1是因为第一列是表头
            sheet.cell(row=row_index, column=col_index).value = item

        wb.save(template_filepath)
        return True
    except Exception as e:
        print(f"复制数据到 Excel 时发生错误：{e}")
        return False


'''
import modules.excel_utils as excel_utils # 推荐使用这种方式，更清晰

def copy_business_numbers_to_template(df, template_filepath):
    """将筛选后的“业务号码”复制到模板文件，分列处理，保留原有表头。"""
    try:
        # 1. 精确剔除开头的 ' 并转换为字符串列表
        df['业务号码'] = df['业务号码'].astype(str).str.replace("^'", "", regex=True) # 使用replace方法，regex=True表示使用正则表达式
        business_numbers = df['业务号码'].tolist()

        print(f"business_numbers 类型：{type(business_numbers)}")
        print(f"business_numbers 内容（前 10 个）：{business_numbers[:10]}")

        # 2. 创建 output_data 二维列表
        max_rows_per_column = 5000
        num_numbers = len(business_numbers)
        num_cols_needed = (num_numbers + max_rows_per_column - 1) // max_rows_per_column
        output_data = [["" for _ in range(num_cols_needed)] for _ in range(max_rows_per_column)]

        for i, number in enumerate(business_numbers):
            row_index = i % max_rows_per_column
            col_index = i // max_rows_per_column
            output_data[row_index][col_index] = str(number)

        print(f"output_data 类型：{type(output_data)}")
        print(f"output_data 内容：{output_data}")

        # 3. 复制数据到 Excel
        header_name = "业务号码"
        if not excel_utils.copy_data_to_excel_with_header(output_data, template_filepath, "文件名称", header_name):
            print("复制业务号码失败")
            return False

        print("业务号码已成功复制到 导入模板.xlsx")
        return True

    except KeyError:
        print("错误：筛选后的数据中缺少 '业务号码' 列。")
        return False
    except Exception as e:
        print(f"处理业务号码或写入“导入模板.xlsx”文件时发生错误：{e}")
        return False


'''

'''
def copy_business_numbers_to_template(df, template_filepath):
    """将筛选后的“业务号码”复制到模板文件，分列处理，保留原有表头，先清空数据。"""
    try:
        business_numbers = df['业务号码'].astype(str).str.lstrip("'").tolist()

        print(f"business_numbers 类型：{type(business_numbers)}")  # 打印类型
        print(f"business_numbers 内容（前 10 个）：{business_numbers[:10]}")  # 打印前 10 个元素

        try:
            template_wb = openpyxl.load_workbook(template_filepath)
            template_sheet = template_wb["文件名称"]
            num_cols = template_sheet.max_column
        except FileNotFoundError:
            print(f"模板文件 {template_filepath} 未找到。")
            return False
        except KeyError as e:
            print(f"模板文件缺少指定sheet：{e}。")
            return False
        except Exception as e:
            print(f"读取模板文件时发生未知错误：{e}")
            return False

        # max_rows_per_column = 5000
        # output_data = []
        max_rows_per_column = 5000
        num_numbers = len(business_numbers)
        num_cols_needed = (num_numbers + max_rows_per_column - 1) // max_rows_per_column  # 计算需要的列数

        output_data = [["" for _ in range(num_cols_needed)] for _ in range(max_rows_per_column)]  # 使用列表推导式创建二维列表

        # for i, number in enumerate(business_numbers):
        #     row_index = i % max_rows_per_column
        #     col_index = i // max_rows_per_column

        print(f"output_data 类型：{type(output_data)}")
        print(f"output_data 内容：{output_data}")

        for i, number in enumerate(business_numbers):
            row_index = i % max_rows_per_column
            col_index = i // max_rows_per_column
            output_data[row_index][col_index] = str(number)  # 强制转换为字符串

            if col_index >= num_cols:
                print("导入模板的列数不足，请添加更多的列。")
                return False

            # while len(output_data) <= row_index:
            #     output_data.append([])
            # while len(output_data[row_index]) <= col_index:
            #     output_data[row_index].append("")
            # output_data[row_index][col_index] = number
            # 正确创建二维列表，确保每个元素都是字符串

            # if len(output_data) <= row_index:
            #     output_data.extend([[""] for _ in range(row_index - len(output_data) + 1)])  # 扩展行，并用空字符串填充
            # while len(output_data[row_index]) <= col_index:
            #     output_data[row_index].append("")  # 扩展列，并用空字符串填充
            # output_data[row_index][col_index] = str(number)  # 强制转换为字符串


        if output_data:
            header_name = "业务号码" # 定义表头名称
            #if not excel_utils.copy_data_to_excel_with_header(output_data, template_filepath, "文件名称",header_name): # 使用带表头的函数，并传递表头名称
            if not excel_utils.copy_data_to_excel_with_header(output_data, template_filepath, "文件名称", "业务号码"):
                print("复制业务号码失败")
                return False

        print("业务号码已成功复制到 导入模板.xlsx")
        return True

    except KeyError:
        print("错误：筛选后的数据中缺少 '业务号码' 列。")
        return False
    except Exception as e:
        print(f"处理业务号码或写入“导入模板.xlsx”文件时发生错误：{e}")
        return False
'''

'''
def copy_business_numbers_to_template(df, template_filepath):
    """将筛选后的“业务号码”复制到模板文件，分列处理，保留原有表头，先清空数据。"""
    try:
        business_numbers = df['业务号码'].astype(str).str.lstrip("'").tolist()

        try:
            template_wb = openpyxl.load_workbook(template_filepath)
            template_sheet = template_wb["文件名称"]
            num_cols = template_sheet.max_column
            #template_wb.close()  # *** 这一行必须删除 ***
        except FileNotFoundError:
            print(f"模板文件 {template_filepath} 未找到。")
            return False
        except KeyError as e:
            print(f"模板文件缺少指定sheet：{e}。")
            return False
        except Exception as e:
            print(f"读取模板文件时发生未知错误：{e}")
            return False

        max_rows_per_column = 5000
        output_data = []

        for i, number in enumerate(business_numbers):
            row_index = i % max_rows_per_column
            col_index = i // max_rows_per_column

            if col_index >= num_cols:
                print("导入模板的列数不足，请添加更多的列。")
                return False

            while len(output_data) <= row_index:
                output_data.append([])
            while len(output_data[row_index]) <= col_index:
                output_data[row_index].append("")
            output_data[row_index][col_index] = number

        if output_data:
            if not excel_utils.copy_data_to_excel(output_data, template_filepath, "文件名称"):
                print("复制业务号码失败")
                return False

        print("业务号码已成功复制到 导入模板.xlsx")
        return True

    except KeyError:
        print("错误：筛选后的数据中缺少 '业务号码' 列。")
        return False
    except Exception as e:
        print(f"处理业务号码或写入“导入模板.xlsx”文件时发生错误：{e}")
        return False
        
'''