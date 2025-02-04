#专门处理 筛选过后的 的 业务号码 数据复制到 导入模板文件

#import pandas as pd
#from modules.excel_utils import copy_data_to_excel

from modules import excel_utils
import openpyxl



from modules import excel_utils
import openpyxl

def copy_business_numbers_to_template(df, template_filepath):
    try:
        business_numbers = df['业务号码'].astype(str).str.replace("'", "", regex=False).tolist()
        print("business_numbers:", business_numbers)

        try:
            template_wb = openpyxl.load_workbook(template_filepath)
            template_sheet = template_wb["文件名称"]
            num_cols = template_sheet.max_column
            if num_cols < 1:  # 关键修改：确保 num_cols 至少为 1
                num_cols = 1
        except FileNotFoundError:
            print(f"模板文件 {template_filepath} 未找到。")
            return False
        except KeyError as e:
            print(f"模板文件缺少指定sheet：{e}。")
            return False
        except Exception as e:
            print(f"读取模板文件时发生未知错误：{e}")
            return False

        max_rows_per_column = 5000
        all_rows = []

        for i, number in enumerate(business_numbers):
            row_index = i % max_rows_per_column
            col_index = i // max_rows_per_column

            while len(all_rows) <= row_index:
                all_rows.append([""] * num_cols)

            all_rows[row_index][col_index] = number

        print("all_rows:", all_rows) # 打印 all_rows

        # if all_rows:
        #     if not excel_utils.copy_data_to_excel(all_rows, template_filepath, "文件名称"):
        #         print("复制业务号码失败")
        #         return False
        if all_rows:
            if not excel_utils.copy_data_to_excel(all_rows, template_filepath, "文件名称"):  # 传递 all_rows
                print("复制业务号码失败")
                return False

        print("业务号码已成功复制到 导入模板.xlsx")
        return True

    except KeyError:
        print("错误：筛选后的数据中缺少 '业务号码' 列。")
        return False
    except Exception as e:
        print(f"处理业务号码或写入“导入模板.xlsx”文件时发生错误：{e}")
        return False

# from modules import excel_utils
# import openpyxl
#
# def copy_business_numbers_to_template(df, template_filepath):
#     try:
#         # 使用 astype(str) 确保所有值都是字符串，然后移除单引号
#         #business_numbers = df['业务号码'].astype(str).str.replace("'", "", regex=False).tolist() # 使用replace移除单引号
#         business_numbers = df['业务号码'].astype(str).str.replace("'", "", regex=False).tolist()
#         print("business_numbers:", business_numbers)  # 打印 business_numbers
#
#         try:
#             template_wb = openpyxl.load_workbook(template_filepath)
#             template_sheet = template_wb["文件名称"]
#             num_cols = template_sheet.max_column
#         except FileNotFoundError:
#             print(f"模板文件 {template_filepath} 未找到。")
#             return False
#         except KeyError as e:
#             print(f"模板文件缺少指定sheet：{e}。")
#             return False
#         except Exception as e:
#             print(f"读取模板文件时发生未知错误：{e}")
#             return False
#
#         max_rows_per_column = 5000
#         all_rows = []
#
#         for i, number in enumerate(business_numbers):
#             row_index = i % max_rows_per_column
#             col_index = i // max_rows_per_column
#
#             # 确保 all_rows 有足够的行
#             while len(all_rows) <= row_index:
#                 all_rows.append([""] * num_cols)
#
#             all_rows[row_index][col_index] = number # 正确赋值
#
#         print("all_rows:", all_rows)  # 打印 all_rows
#
#         if all_rows:
#             if not excel_utils.copy_data_to_excel(all_rows, template_filepath, "文件名称"):
#                 print("复制业务号码失败")
#                 return False
#
#         print("业务号码已成功复制到 导入模板.xlsx")
#         return True
#
#     except KeyError:
#         print("错误：筛选后的数据中缺少 '业务号码' 列。")
#         return False
#     except Exception as e:
#         print(f"处理业务号码或写入“导入模板.xlsx”文件时发生错误：{e}")
#         return False

# def copy_business_numbers_to_template(df, template_filepath):
#     """将筛选后的“业务号码”复制到模板文件，分列处理，保留原有表头，先清空数据。"""
#     try:
#         business_numbers = df['业务号码'].astype(str).str.lstrip("'").tolist()
#     except KeyError:
#         print("错误：筛选后的数据中缺少 '业务号码' 列。")
#         return False
#
#         # try:
#         #     template_wb = openpyxl.load_workbook(template_filepath)
#         #     template_sheet = template_wb["文件名称"]
#         #     num_cols = template_sheet.max_column
#         #     #template_wb.close()  # *** 删除这一行 否则无法写入模板文件***
#         # except FileNotFoundError:
#         #     print(f"模板文件 {template_filepath} 未找到。")
#         #     return False
#         # except KeyError as e:
#         #     print(f"模板文件缺少指定sheet：{e}。")
#         #     return False
#         # except Exception as e:
#         #     print(f"读取模板文件时发生未知错误：{e}")
#         #     return False
#
#         max_rows_per_column = 5000
#         num_cols = template_sheet.max_column  # 确保在这里获取num_cols
#         #output_data = []
#         all_rows = []  # 用于存储所有行的数据
#         for i, number in enumerate(business_numbers):
#             row_index = i % max_rows_per_column
#             col_index = i // max_rows_per_column
#
#             # # 如果是新的一列，则添加新的空行
#             # if col_index >= len(all_rows):
#             #     for _ in range(max_rows_per_column):
#             #         all_rows.append([""] * num_cols)  # 初始化所有单元格为空字符串
#             # 确保all_rows有足够的行
#             while len(all_rows) <= row_index:
#                 all_rows.append([""] * num_cols)  # 初始化行为空字符串列表
#
#             all_rows[row_index][col_index] = number  # 直接按行列赋值
#
#         if all_rows:
#             if not excel_utils.copy_data_to_excel(all_rows, template_filepath, "文件名称"):  # 传递all_rows
#                 print("复制业务号码失败")
#                 return False
#
#         print("业务号码已成功复制到 导入模板.xlsx")
#         return True
#
#         # for i, number in enumerate(business_numbers):
#         #     row_index = i % max_rows_per_column
#         #     col_index = i // max_rows_per_column
#         #
#         #     if col_index >= num_cols:
#         #         print("导入模板的列数不足，请添加更多的列。")
#         #         return False
#         #
#         #     while len(output_data) <= row_index:
#         #         output_data.append([])
#         #     while len(output_data[row_index]) <= col_index:
#         #         output_data[row_index].append("")
#         #     output_data[row_index][col_index] = number
#         #
#         # if output_data:
#         #     if not excel_utils.copy_data_to_excel(output_data, template_filepath, "文件名称"):
#         #         print("复制业务号码失败")
#         #         return False
#         #
#         # print("业务号码已成功复制到 导入模板.xlsx")
#         # return True
#
#
#     except KeyError:
#         print("错误：筛选后的数据中缺少 '业务号码' 列。")
#         return False
#     except Exception as e:
#         print(f"处理业务号码或写入“导入模板.xlsx”文件时发生错误：{e}")
#         return False



#原本逻辑 没有实现
'''
def copy_business_numbers_to_template(df, template_filepath):
    """将筛选后的“业务号码”复制到模板文件"""
    try:
        business_numbers = df['业务号码'].astype(str).str.lstrip("'")
        if not copy_data_to_excel(business_numbers, template_filepath, "Sheet1", "业务号码"):
            print("复制业务号码失败")
            return False
        print("业务号码已成功复制到 导入模板.xlsx")
        return True
    except KeyError:
        print("错误：筛选后的数据中缺少 '业务号码' 列。")
        return False
    except Exception as e:
        print(f"处理业务号码或写入“导入模板.xlsx”文件时发生错误：{e}")
        return False
'''